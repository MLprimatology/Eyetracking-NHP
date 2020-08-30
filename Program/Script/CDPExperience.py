# -*- coding: utf8 -*-

# La classe CDPexperience hérite de la classe CDPApplication du fichier python CDPInterface
#  Elle implémente les méthodes virtuelles héritées de la classe mère




__authors__ = ("Mathieu Legrand") 
__contact__ = ("mathieu.legrand78@gmail .com")
__version__ = "1.0.1"
__copyright__ = "copyleft" 
__date__ = "28/08/2020"


import random
import os
import time
import datetime
import json

import pygaze

from openpyxl import Workbook,load_workbook


import tobii_research as tr

from os.path import basename
from functools import partial
from pygaze import libscreen
from pygaze import libtime
from pygaze import liblog
from pygaze import libinput
from pygaze import libsound
from Visualisation import *
from pygaze.eyetracker import EyeTracker
from cdplibtobii import CDPProTracker
from CDPInterface import CDPApplication
import tkinter, Tkconstants, tkFileDialog, tkMessageBox
from pygaze.keyboard import Keyboard
from tkinter import *
import tkinter.messagebox



import xlrd 





 
class experience(CDPApplication):
	def __init__(self,title):
		super(experience,self ).__init__(title)

	# Calibration de l'eye tracker 

	def CDPInitialisation(self):
		self.findtracker =  tr.find_all_eyetrackers()
		if self.findtracker == ():
			print("Veuillez réassayer, aucun EyeTracker détecté")
			return()
		self.filename = 0
		self.baseFilename = 0
		self.blackdisp = libscreen.Display(screennr=int(self.Config.getConfiguration('DISPLAY','screen_number')))

		self.disp = libscreen.Display(screennr=int(self.Config.getConfiguration('DISPLAY','screen_number')))
		self.blankscreen = libscreen.Screen()

		self.tracker = CDPProTracker(self.disp)# création de l'objet eyeTracker
		self.kb = Keyboard(keylist=['space', 'escape', 'q'], timeout=1)
		self.Visu = CDPBaseVisualisation(self)
		self.RecSound = libsound.Sound(soundfile=self.Config.getSoundDirname ('2.wav'))
		self.ErrSound = libsound.Sound(soundfile=self.Config.getSoundDirname ('punition.wav'))
		self.nameInd = 0
		self.mydisp = [self.disp]


		print("Eyetracker connecté avec succès")

	def CDPDisconnect(self):
		if self.tracker == 0:
			print("Aucun Eyetracker connecté")
			return()
		self.tracker.close()
		self.disp.close()

		print("Eyetracker déconnecté avec succès")



	def CDPinitCalibration(self):
		self.AffichageVis()
		self.filename = tkFileDialog.asksaveasfilename(initialdir = self.Config.getCalibrationDir(),title = "Select file",filetypes = (("calibration files","*.bin"),("all files","*.*"))) 
		self.baseFilename = os.path.splitext(self.filename)[0]  
		self.tracker.initCalibration(self.Visu)
		self.nameInd = basename(self.baseFilename)
		self.Visu.delete()
		self.disp.close()



	def CDPSendPoint(self, num):
		

		self.tracker.addCalibrationPoint( num-1 )




	def CDPValidatePoint(self,num):

		time.sleep(0.8)
		self.tracker.validateCalibrationPoint(num-1)
		self.RecSound.play()
        
		
	def CDPApplyCalibration(self):
		        

		ListeCal = self.tracker.applyCalibration(self.baseFilename)
		self.disp = libscreen.Display(screennr=int(self.Config.getConfiguration('DISPLAY','screen_number')))

		return(ListeCal)
		
	def CDPEndCalibration(self,Liste):

		date = datetime.datetime.now().strftime("%d-%m-%Y_%H:%M:%S")
		self.saveData(self.baseFilename +'_' + date +'.bin')
		table = Workbook()
		gazePosSheet = table.active
		gazePosSheet.title = 'gazePos'
		gazePosSheet.append(["Xref","Yref", "XOeilDroit", "YOeilDroit","XOeilGauche","YOeilGauche"])
		for Lis in Liste:
			gazePosSheet.append(Lis)

		table.save(self.baseFilename +'_' +  date  +'.xls' )
		self.baseFilename = self.baseFilename +'_' + date

		self.root.title("Fenêtre de suivi de " + self.nameInd )

	def CDPSendValidationPoint(self,indice):

		self.tracker.showPoint(indice)

	def CDPeyeData(self,indice):
		Liste = self.tracker.RecupDonneesOeil(indice)
		self.RecSound.play()
		return(Liste)



	def CDPSaveGain(self,ListeOeilValid):
		table = Workbook()
		gazePosSheet = table.active
		gazePosSheet.title = 'gazePos'
		gazePosSheet.append(["Xref","Yref", "XOeilDroit", "YOeilDroit","XOeilGauche","YOeilGauche","xRetenu","Yretenu"])
		for Lis in ListeOeilValid:
			for L in Lis:
				gazePosSheet.append([L[0][0],L[0][1],L[1][0],L[1][1],L[2][0],L[2][1],L[3][0],L[3][1]])
		table.save(self.Config.getCalibrationDir() +'/' + str(self.nameInd) +'_Validation_calibration.xls' )

	def CDPLoadCalibration(self):

		'''
		Descirption : 
			Permet le chargement d'un fichier de calibration préenregistré sur l'ordinateur et de l'appliquer à l'EyeTracker
		'''

		self.filename = tkFileDialog.askopenfilename(initialdir = self.Config.getCalibrationDir(),title = "Select file",filetypes = (("calibration files","*.bin"),("all files","*.*")))
		if not self.filename:
			tkinter.messagebox.showwarning(title= "Attention", message='Pas de fichier de calibration trouve')
		else : 
			with open(self.filename, "rb") as f:
				calibration_data = f.read()
				if len(calibration_data) > 0:
					print("Applying calibration on eye tracker with serial number {0}.".format(self.tracker.eyetracker.serial_number))
					self.tracker.eyetracker.apply_calibration_data(calibration_data)
			self.baseFilename = os.path.splitext(self.filename)[0]
			self.nameInd = basename(self.baseFilename)
		self.root.title("Fenêtre de suivi de " + self.nameInd )

	def CDPLeaveCalibration(self):
		self.tracker.annuler_calibration()
		self.disp = libscreen.Display(screennr=int(self.Config.getConfiguration('DISPLAY','screen_number')))

	def CDPVisualisationCalibration(self):
		if self.baseFilename == 0:
			tkinter.messagebox.showerror(title= "Erreur", message='Veuillez selectionner un fichier de calibration')
		else :
			myBook = xlrd.open_workbook(self.baseFilename+ '.xls')
			ControlSHeet = myBook.sheet_by_index(0) 
			nbRow = ControlSHeet.nrows
			Liste = []
			for i in range (1,nbRow):
				Liste += [ControlSHeet.row_values(i)]
			for i in range (len(Liste)):
				for j in range (len(Liste[i])):
					if Liste[i][j] == u'' :
						Liste[i][j] = -1


			self.tkplot = tk.Toplevel(self.root)
			self.tkplot.title ("Représentation Calibration") # nom de la fenetre tkinter fille
			self.fig = Figure()

			self.canvas = FigureCanvasTkAgg(self.fig, master=self.tkplot)  
			self.canvas.get_tk_widget().pack(fill='both', expand=True)
			ax = self.fig.add_subplot(111)
			ax.set_ylim(0,1080)
			ax.invert_yaxis()
			ax.set_xlim(0,1920)
			
			col = ['black','grey','yellow','green','pink','red','blue','purple','orange']
			cpt = 0


			OldXref = Liste[0][0]
			OldYref = Liste[0][1]
			ListeXref = [OldXref]
			ListeYref = [OldYref]
			ListeTest = [(OldXref,OldYref)]
			for Pos in Liste:
				Xref = Pos[0]
				Yref = Pos[1]
				if (Xref,Yref) not in ListeTest:
					ListeXref += [Xref]
					ListeYref += [Yref]
					ListeTest += [(Xref,Yref)]
					cpt +=1

				if Pos[2] != 0 and Pos[3] != 0:
					ax.scatter(Pos[2],Pos[3], c = col[cpt],marker='x',s=8)
				if Pos[4] != 0 and Pos[5] != 0:
					ax.scatter(Pos[4],Pos[5],c=col[cpt],marker='s',s=8)
			for i in range (len (ListeXref)):
				ax.scatter(ListeXref[i],ListeYref[i],c=col[i],s=80)

			self.canvas.get_tk_widget().pack(fill='both', expand=True)
			self.canvas.show()	



	def CDPApprentissagePos(self,Name,Cond,Rec):

		dirname = self.Config.getDataDirname('ValidationPosition') + '/'
		date = datetime.datetime.now().strftime("%d-%m-%Y_%H:%M:%S")
		jour = datetime.datetime.now().strftime("%d/%m/%Y")
		self.tracker.start_recording()
		Liste,tablepos = self.tracker.ApprentissagePos(self.Visu)
		self.tracker.stop_recording()
		Liste[0] = Name
		Liste[1] = date
		Liste[2] = jour
		Liste[8] = Cond
		Liste[9] = Rec
		Liste[12] = self.remarque.get()
		table = load_workbook(dirname + 'ApprentissagePosition.xlsx')
		mysheet = table.active
		mysheet.append(Liste)
		table.save(dirname + 'ApprentissagePosition.xlsx')
		tablepos.save(dirname + Name + '_' + date +'.xlsx' )
		self.disp.fill(screen= self.blankscreen )
		self.disp.show()




########## Gestion de l interface tkinter ###############

### Gestion des fichiers textes ###

	'''
	
	Description : Les trois fonctions suivantes permettents successivement :
		- L'ouverture d'un fichier texte avec la date
		- L'inscription des données dans ce fichier 
		- La fermeture de ce fichier texte

    '''



	def LogData(self,t,x,y,Rx,Ry):
		self.file.write("%d;%d;%d;%d;%d" %(t,x,y,Rx,Ry))
		self.file.write('\n')

	
	def saveData (self,filename):

		'''
		Paramètres : 
			filename : Nom du fichier de sauvegarde, en général nom de l'individu présent devant l'EyeTracker 

		Description :
			Permet la sauvegarde du fichier .bin de la calibration pour la réutiliser
		'''

		with open(filename, "wb") as f:
			calibration_data = self.tracker.eyetracker.retrieve_calibration_data()
			if calibration_data is not None:
				print("Saving calibration to file for eye tracker with serial number {0}.".format(self.tracker.eyetracker.serial_number)) 
				f.write(self.tracker.eyetracker.retrieve_calibration_data())
			else:
				print("No calibration available for eye tracker with serial number {0}.".format(self.tracker.eyetracker.serial_number)) 




### Conversion point normalise / pixel 

	def norm_2_px(self,normalized_point):

		'''
		Paramètres : 
			normalized_point : coordonnées d'un point en coordonnées "on display Area" (x et y compris entre 0 et 1)
		Description :
			Permet la conversation des coordonnées de ce point en pixel 
		'''
		return (round(normalized_point[0] * self.disp.dispsize[0], 0), round(normalized_point[1] *self.disp.dispsize[1], 0))

	def px_2_norm(self,pixelized_point):

		'''
		
		Paramètres : 
			normalized_point : coordonnées d'un point en coordonnées pixelisé
		Description :
			Permet la conversation des coordonnées de ce point en coordonnées normales (comprises entre 0 et 1) 

		'''

		return (pixelized_point[0] / self.disp.dispsize[0], pixelized_point[1] / self.disp.dispsize[1])

	def CDPeffacer(self):

		'''
		Description :
			Permet d'effacer l'affichage de l'écran présenté et de remettre un écran noir  
		'''
		self.disp.fill(screen=self.blankscreen)
		self.disp.show()

### Experience mouvements de balle

	def showPoint( self,point):

		'''
		Paramètre:
			point : coordonées du point en pixel

		Description :
			Permet l'affichage d'un point jaune aux coordonnées entrées 
		'''

		self.screen.clear()
        # CDP : Changement couleur
		self.screen.draw_circle(colour='yellow', pos=point, r=30, pw=5, fill=True)
		self.disp.fill(self.screen)
		self.disp.show()


	def CDPaffichageVis(self):

		'''
		Description :
			Permet l'afficahge d'un écran de contrôle sur lécran de l'éxpérimentateur
		'''

		self.Visu = CDPVisualisation(self,self.disp)

	def CDPDeleteVisualisation(self):
		self.Visu = CDPBaseVisualisation(self)


###### Self_Guided_Visual_Session  -   SGVS ###########


	def ValidationFixation(self,screen,t):

		'''
		Paramètres :
			screen : écran virtuel d'affichage
			t : temps de fixation en millisecondes

		Description :
			Permet d'afficher un point sur l'écran qui devient vert lorsque l'on regarde ce point et attendre de le fixer pendant "t" millisecondes. 

		'''	

		tfix = 0
		

		while tfix < t: # Tant que l'on regarde pas le cercle pendant 1 seconde 

			newTime = libtime.get_time()
			gazepos = self.tracker.sample()
			# si l'individu ne regarde pas le point central

			if (gazepos[0] < self.norm_2_px((0.44,0.44))[0] or gazepos[0] > self.norm_2_px((0.56,0.56))[0]) and (gazepos[1] < self.norm_2_px((0.44,0.44))[1] or gazepos[1] > self.norm_2_px((0.56,0.56))[1]) : 
				screen = libscreen.Screen()
				#le cercle reste blanc
				screen.draw_circle(colour='white', pos= self.norm_2_px((0.5,0.5)), r=40, pw=2, fill=True) 
				self.disp.fill(screen=screen)
				self.disp.show()
				tfix = 0

			else : # si l'individu regarde le point central
				screen = libscreen.Screen()
				#Le cercle devient vert
				screen.draw_circle(colour='green', pos= self.norm_2_px((0.5,0.5)), r=40, pw=2, fill=True)
				self.disp.fill(screen=screen)
				self.disp.show()
				tfix += (libtime.get_time() - newTime)
				

			if self.kb.get_key(keylist= ['space'], flush=False)[0]:
				screen.clear()
				self.disp.fill(screen=screen)
				self.disp.show()

				return()
		screen.clear()




	def CDPSGVS(self):

		'''
		Description:
			Après avoir fixé un point central pendant un certains temps (convenu dans le fichier .json), 4 formes apparaissent sur l'écran (rond, carré, triangle et losange)
			Les résulats de l'expérience sont sauvegardées dans un fichier excel
		'''



		if self.filename == 0 :
			tkinter.messagebox.showerror(title= "Erreur", message='Veuillez selectionner un fichier de calibration')
		else :

			date = datetime.datetime.now().strftime("%d-%m-%Y_%H:%M:%S")
			# Lecture des paramètres de l'expérience
			filename = self.Config.getConfigFilename('SGVS')
			try:
				with open(filename , "r") as fichier:

					data = json.load(fichier)

					CheminSave = self.Config.getDataDirname (str(data["result_dirname"])) + '/'
					
					duration = data["duration"] 
					nbTrial = data["nb_trial"]
					color = data["color"]
					fixationStart = data["fixation_start"] 
					frequency = data["frequency"]

			except IOError :
				print ("fichier %s introuvable", filename )

	
			date = datetime.datetime.now().strftime("%d-%m-%Y_%H:%M:%S")

			# Creation du fichier excel de sauvegarde des résultats
			table = Workbook()
			gazePosSheet = table.active
			gazePosSheet.title = 'gazePos'
			posStimuli = table.create_sheet("posStimuli")

			gazePosSheet.append(["Time", "XOeilDroit", "YOeilDroit","XOeilGauche","YOeilGauche","xRetenu","Yretenu","Etat"])
			posStimuli.append(["Id Stimuli","Temps apparition Stimuli","Position X Stiumli","Position y Stiumli","Temps Disparition Stimuli"])

			screenSGVS = libscreen.Screen()
			NbrDiag = 0  ### Nombre de fois d apparition des stimulis en diagonales
			NbrDroits = 0 ##" nombre de fois d apparition des stimulis en "croix droite"
			self.tracker.start_recording()

			PositionDiag = [(0.1,0.1),(0.9,0.1),(0.1,0.9),(0.9,0.9)] ### Posiitions possible des 4 stimulis en croix diagonale
			PositionDroit = [(0.5,0.1),(0.9,0.5),(0.5,0.9),(0.1,0.5)] ### Posiitions possible des 4 stimulis en croix droite 
			# Boucle sur le nombre d'essais
			for i in range (nbTrial*2):

				self.ValidationFixation(screenSGVS,fixationStart)

				a = random.randint(0,1) ### permet de choisir si les stimuli vont etre choisis en diag ou droit
				PosRetenues=[]

				if NbrDiag == nbTrial:
					a = 1
				if NbrDroits ==nbTrial:
					a = 0

				# Choix aléatoire d'une des 2 listes et mélange des positions des formes
				if a == 0 :
					NbrDiag +=1
					PosRetenues= PositionDiag

				else :
					NbrDroits += 1
					PosRetenues= PositionDroit


				random.shuffle (PosRetenues)
				
				PosStimuliA = self.norm_2_px(PosRetenues[0])
				PosStimuliB = self.norm_2_px(PosRetenues[1])
				PosStimuliC = self.norm_2_px(PosRetenues[2])
				PosStimuliD = self.norm_2_px(PosRetenues[3])

				screenSGVS.draw_rect(x = PosStimuliA[0], y =PosStimuliA[1],w=80,h=80,fill = True, colour=color)	## dessin d'un carré
				screenSGVS.draw_circle(colour=color, pos=PosStimuliB, r=40, pw=5, fill=True) ## dessin d'un cercle
				screenSGVS.draw_polygon([(PosStimuliC[0] -40,PosStimuliC[1]+40),(PosStimuliC[0]+40,PosStimuliC[1]+40),(PosStimuliC[0],PosStimuliC[1]-40)],fill = True, colour=color ) 	## dessin d'un triangle
				screenSGVS.draw_polygon([(PosStimuliD[0] ,PosStimuliD[1]-60),(PosStimuliD[0]-30,PosStimuliD[1]),(PosStimuliD[0],PosStimuliD[1]+60),(PosStimuliD[0]+30,PosStimuliD[1])],fill = True, colour=color ) 	## dessin d'un losange
				
				TempsApp = libtime.get_time()
				self.disp.fill(screen=screenSGVS)
				self.disp.show()

				tdeb = libtime.get_time()
				# Lancement de l'échantillonage / capture de la position des yeux
				oldTimeStamp = 0
				txp,gase = self.tracker.binocular_sample()
				while libtime.get_time()-tdeb < duration :
					time.sleep(frequency)
					NewTimeStamp, Newgazepos = self.tracker.binocular_sample()
					if NewTimeStamp != oldTimeStamp :
						t = int(NewTimeStamp - txp) /1000
						etat = self.etat_yeux(Newgazepos[0],Newgazepos[1])
						gazePosSheet.append([t, Newgazepos[0][0],Newgazepos[0][1],Newgazepos[1][0],Newgazepos[1][1],Newgazepos[2][0],Newgazepos[2][1],etat])
						oldTimeStamp = NewTimeStamp


				screenSGVS.clear()
				self.disp.fill(screen=screenSGVS)
				self.disp.show()
				TempsDisp = libtime.get_time()

				time.sleep(2)	
				
				# Ajoute la coordonnées des formes dans l'excel

				posStimuli.append(["1",TempsApp,PosStimuliA[0],PosStimuliA[1],TempsDisp]) # le 1 représente le carré
				posStimuli.append(["2",TempsApp,PosStimuliB[0],PosStimuliB[1],TempsDisp]) # le 2 représente le cercle
				posStimuli.append(["3",TempsApp,PosStimuliC[0],PosStimuliC[1],TempsDisp]) # le 3 représente le triangle
				posStimuli.append(["4",TempsApp,PosStimuliD[0],PosStimuliD[1],TempsDisp]) # le 4 représente le losange

			self.tracker.stop_recording()
			table.save(CheminSave + self.nameInd + '_' + date +  '_Donnees.xls')
			tkinter.messagebox.showinfo("Fichier sauvegarde","Le fichier a bien été sauvegardé")


#### Fonction FixationVisage ######


	def fonction_essai(self):
		
		ScreenVisage = libscreen.Screen()
		self.tracker.start_recording()
		ListeNomImg = ['/home/eyetracker/Bureau/Program/Images/pizza.jpg','/home/eyetracker/Bureau/Program/Images/ExpVisages/Image_Homme_2.bmp']
		cpt = 0


		for Img in ListeNomImg:

			table = Workbook()
			gazePosSheet = table.active
			gazePosSheet.title = 'gazePos'

			gazePosSheet.append(["Time", "XOeilDroit", "YOeilDroit","XOeilGauche","YOeilGauche","xRetenu","Yretenu","Etat"])
			self.ValidationFixation(ScreenVisage,1000)

			LisTOeilDroit  = []
			LisTOeilGauche = []
			ListeOeilRetenu = []

			temps = []
			ScreenVisage.draw_image(image =Img)
			self.disp.fill(screen=ScreenVisage)
			self.disp.show()
			tdeb = libtime.get_time()

			oldTimeStamp = 0
			while libtime.get_time()-tdeb < 10000 :
				time.sleep(0.010)
				t = libtime.get_time()
				NewTimeStamp, Newgazepos = self.tracker.binocular_sample()
				if NewTimeStamp != oldTimeStamp :
					etat = self.etat_yeux(Newgazepos[0],Newgazepos[1])
					gazePosSheet.append([t, Newgazepos[0][0],Newgazepos[0][1],Newgazepos[1][0],Newgazepos[1][1],Newgazepos[2][0],Newgazepos[2][1],etat])
					oldTimeStamp = NewTimeStamp


			
			ScreenVisage.clear()
			self.disp.fill(screen=ScreenVisage)
			self.disp.show()
			table.save('/home/eyetracker/Bureau/'  + self.nameInd + '_'+ str(cpt) +  '_Donnees.xls')
			cpt = 1
		self.tracker.stop_recording()

	def CDPFixVisage(self):	

		'''
		Description:
			Après avoir fixé un point pendant un certain (défini dans le fichier .json), un visage apparaît à droite ou à gauche de l'écran. L'individu peut alors regarder l'image pendant 5 secondes.
			Il y a 8 images de femmes, 8 images d'hommes et 8 images de Tonkeans. 
			Des images sont sauvegardées avec une heatmap des temps de fixation de l'individu, des saccades et des points où il a regardé. 
		'''
	
		if self.filename == 0 :
			tkinter.messagebox.showerror(title= "Erreur", message='Veuillez selectionner un fichier de calibration')
		else :


			# Lecture des paramètres de l'expérience

			filename = self.Config.getConfigFilename('VISAGES')
			try:
				with open(filename , "r") as fichier:
					data = json.load(fichier)
					dirImage =  self.Config.getImageDirname (str(data["image_dirname"])) +'/'
					CheminSave = self.Config.getDataDirname (str(data["result_dirname"])) + '/'
					duration = data["duration"] 
					fixationStart = data["fixation_start"] 
					frequency = data["frequency"] 

			except IOError :
				print ("fichier %s introuvable", filename )


			date = datetime.datetime.now().strftime("%d-%m-%Y_%H:%M:%S")
			os.mkdir(CheminSave + self.nameInd + '_' + date)

			dirToSave = CheminSave + self.nameInd + '_' + date + '/'

			ScreenVisage = libscreen.Screen()
			# Liste des noms d'images à afficher

			ListeNomImg = os.listdir(dirImage)


			#mélange aléatoirement la liste pour afficher les images dans un ordre aléatoire
			random.shuffle (ListeNomImg)

			
			self.tracker.start_recording()
			for Img in ListeNomImg:
				table = Workbook()
				gazePosSheet = table.active
				gazePosSheet.title = 'gazePos'
				informationsheet = table.create_sheet("Information")


				gazePosSheet.append(["Time", "XOeilDroit", "YOeilDroit","XOeilGauche","YOeilGauche","xRetenu","Yretenu","Etat"])
				informationsheet.append(["NomImg","xImage","yImage"])

				self.ValidationFixation(ScreenVisage,fixationStart)

				a = random.randint(0,1)

				# Choix aléatoire de la position de l'image (à droite ou à gauche)

				if a == 0 :
					Pos = self.norm_2_px((0.25,0.5))
					
				else : 
					Pos = self.norm_2_px((0.75,0.5))


				LisTOeilDroit  = []
				LisTOeilGauche = []
				ListeOeilRetenu = []

				temps = []

				ScreenVisage.draw_image(image = dirImage + Img, pos=Pos)
				self.disp.fill(screen=ScreenVisage)
				self.Visu.affiche_image(dirImage + Img,Pos)
				self.disp.show()
				tdeb = libtime.get_time()
				oldTimeStamp = 0
				txp,gase = self.tracker.binocular_sample()
				while libtime.get_time()-tdeb < duration :
					time.sleep(frequency)
					NewTimeStamp, Newgazepos = self.tracker.binocular_sample()
					self.Visu.Show_gaze(Newgazepos[2][0],Newgazepos[2][1])
					if NewTimeStamp != oldTimeStamp :
						t = int(NewTimeStamp - txp) /1000
						etat = self.etat_yeux(Newgazepos[0],Newgazepos[1])
						gazePosSheet.append([t, Newgazepos[0][0],Newgazepos[0][1],Newgazepos[1][0],Newgazepos[1][1],Newgazepos[2][0],Newgazepos[2][1],etat])
						oldTimeStamp = NewTimeStamp
				ScreenVisage.clear()
				self.disp.fill(screen=ScreenVisage)
				self.disp.show()
				self.Visu.blackscreen()
				informationsheet.append([Img,Pos[0],Pos[1]])

				table.save(dirToSave  + self.nameInd + '_'+ os.path.splitext(Img)[0] +  '_Donnees.xls')

			self.tracker.stop_recording()

	def CDPcontroleNictation(self):


		'''
		Description :
			Après avoir fixé un point central pendant 500 ms, une image de mer (paysage calm et apsiant) apparait pednant 30 sec puis un autre point a fixer et une 
			image de "Ou est Charlie" apparait et l'individu doit trouver Charlier (COncentration). Cette foction renvoie un tableau excel avec les données de l'eye tracker
			permettant d'effectuer un contrôle psitiif sur les nictations de l'individu
		'''
		
		date = datetime.datetime.now().strftime("%d-%m-%Y_%H:%M:%S")
		self.tracker.start_recording()
		name = 'Charlie' 
		ScreenNict = libscreen.Screen()
		Img = self.Config.getImageDirname ('ControlNictation') + '/Charlie.jpg'

		table = Workbook()
		gazePosSheet = table.active
		gazePosSheet.title = 'gazePos'
					
		informationsheet = table.create_sheet("Information")
		gazePosSheet.append(["Time", "XOeilDroit", "YOeilDroit","XOeilGauche","YOeilGauche","xRetenu","Yretenu","Etat"])
		informationsheet.append(["NomImg","xImage","yImage"])
		print("Début de l'expérience")
		self.ValidationFixation(ScreenNict,500)

		ScreenNict.draw_image(image = Img)
		self.disp.fill(screen=ScreenNict)
		self.disp.show()


		tdeb = libtime.get_time()
		oldTimeStamp = 0
		txp,gase = self.tracker.binocular_sample()
		while libtime.get_time()-tdeb < 30000 :
			time.sleep(0.01)
			NewTimeStamp, Newgazepos = self.tracker.binocular_sample()
			if NewTimeStamp != oldTimeStamp :
				t = int(NewTimeStamp - txp) /1000
				etat = self.etat_yeux(Newgazepos[0],Newgazepos[1])
				gazePosSheet.append([t, Newgazepos[0][0],Newgazepos[0][1],Newgazepos[1][0],Newgazepos[1][1],Newgazepos[2][0],Newgazepos[2][1],etat])
				oldTimeStamp = NewTimeStamp
			if self.kb.get_key(keylist= ['space'], flush=False)[0]:
				ScreenNict.clear()
				self.disp.fill(screen=ScreenNict)
				self.disp.show()
				self.tracker.stop_recording()
				print("Fin de l'expérience")

				return()

		ScreenNict.clear()
		self.disp.fill(screen=ScreenNict)
		self.disp.show()

		informationsheet.append([name +'.jpg',960,540])

		table.save('/home/eyetracker/Bureau/Data/Experiences/ControlNictation/' +self.nameInd +'_' + name +'_'+ date + '.xls')
		print("Fin de l'expérience")

		self.tracker.stop_recording()
		
	def CDPFixationPoint(self,tfixation,Name,tol,x,y):

		self.Visu.blackscreen()
		datej = datetime.datetime.now().strftime("%d-%m-%Y")
		
		date = datetime.datetime.now().strftime("%d-%m-%Y_%H:%M:%S")

		table = load_workbook(self.Config.getDataDirname('FixationPoint') + '/' + Name + '.xlsx')
		mysheet = table.active
		ListeResultat = []
		txp = 0
		tdeb = 0
		tfix = 0
		cptperte = 0
		self.tracker.start_recording()
		#table = Workbook()
		#gazePosSheet = table.active
		#gazePosSheet.title = 'gazePos'
		Result = 1			
		#gazePosSheet.append(["Time", "XOeilDroit", "YOeilDroit","XOeilGauche","YOeilGauche","xRetenu","Yretenu","Etat"])
		pospxl = self.norm_2_px((x,y))

		self.Visu.draw_AOI_fix(pospxl[0],pospxl[1],tol)
		self.Visu.VisuShow()
		screen = libscreen.Screen()
		cptdefauteyetracker = 0
		#le cercle reste blanc
		screen.draw_circle(colour='white', pos= pospxl, r=30, pw=2, fill=True) 
		self.disp.fill(screen=screen)
		self.disp.show()
		oldTimeStamp = 0
		tdebxp = libtime.get_time()
		txp ,Newgazepos = self.tracker.binocular_sample()
		while tdeb < 3000 and not ((Newgazepos[2][0] > (pospxl[0] - tol) and Newgazepos[2][0] < (pospxl[0] + tol)) and (Newgazepos[2][1] > (pospxl[1] - tol) and Newgazepos[2][1] < (pospxl[1] + tol))):
			print(tdeb)
			time.sleep(0.005)
			NewTimeStamp, Newgazepos = self.tracker.binocular_sample()
			self.Visu.Show_gaze(Newgazepos[2][0],Newgazepos[2][1])
			newTime = libtime.get_time()
			if NewTimeStamp != oldTimeStamp :
				t = int(NewTimeStamp - txp) /1000
				etat = self.etat_yeux(Newgazepos[0],Newgazepos[1])
				#gazePosSheet.append([t, Newgazepos[0][0],Newgazepos[0][1],Newgazepos[1][0],Newgazepos[1][1],Newgazepos[2][0],Newgazepos[2][1],etat])
				oldTimeStamp = NewTimeStamp
			tdeb = (libtime.get_time() - tdebxp)

		screen= libscreen.Screen()

		if tdeb < 3000:
			
			while tfix < tfixation and cptperte<300:  

				newTime = libtime.get_time()
				time.sleep(0.005)
				NewTimeStamp, Newgazepos = self.tracker.binocular_sample()
				self.Visu.Show_gaze(Newgazepos[2][0],Newgazepos[2][1])
				self.Visu.VisuShow()
				#si l'individu ne regarde pas le point central

				if (Newgazepos[2][0] > (pospxl[0] - tol) and Newgazepos[2][0] < (pospxl[0] + tol)) and (Newgazepos[2][1] > (pospxl[1] - tol) and Newgazepos[2][1] < (pospxl[1] + tol)) : 
					screen = libscreen.Screen()
					#Le cercle devient vert
					#screen.draw_circle(colour=(int(128-(tfix*128/tfixation)),128,0), pos= pospxl, r=30, pw=2, fill=True)
					#screen.draw_circle(colour='green', pos= pospxl, r=30, pw=2, fill=True)
					screen.draw_circle(colour='white', pos= pospxl, r=30, pw=2, fill=True)


					self.disp.fill(screen=screen)
					self.disp.show()
					tfix += (libtime.get_time() - newTime)
					cptperte = 0
					cptdefauteyetracker = 0
					
				else : # si l'individu regarde le point central
					if Newgazepos[2] == (-1,-1):
						cptdefauteyetracker += (libtime.get_time() - newTime)
					cptperte += (libtime.get_time() - newTime)



				if NewTimeStamp != oldTimeStamp :
					t = int(NewTimeStamp - txp) /1000
					etat = self.etat_yeux(Newgazepos[0],Newgazepos[1])
					#gazePosSheet.append([t, Newgazepos[0][0],Newgazepos[0][1],Newgazepos[1][0],Newgazepos[1][1],Newgazepos[2][0],Newgazepos[2][1],etat])
					oldTimeStamp = NewTimeStamp

			if tfix> tfixation:
				self.RecSound.play()
				Result = 0
		self.tracker.stop_recording()

		if tdeb>3000:
			Resulat =1 
			self.ErrSound.play()
			screen = libscreen.Screen(bgc='white')
			self.disp.fill(screen=screen)
			self.disp.show()
			time.sleep(3)

		elif tfix <tfixation :
			if cptdefauteyetracker > 200:
				Result = 2
			else :
				Result = 3
			self.ErrSound.play()
			screen = libscreen.Screen(bgc='white')
			self.disp.fill(screen=screen)
			self.disp.show()
			time.sleep(3)

		mysheet.append([datej,x,y,tol,tfixation,Result])

		#table.save(self.Config.getDataDirname('FixationPoint') + '/' + Name +'_' +  date + '_' +  Result + '_' + str(tfixation) + 'ms' + '.xls')
		table.save(self.Config.getDataDirname('FixationPoint') + '/' + Name + '.xlsx')

		print('fin')

		screen = libscreen.Screen()
		self.disp.fill(screen=screen)
		self.disp.show()

	def CDPExplorationVisage(self,name,xfix,yfix):

		xfix = float(xfix)
		yfix = float(yfix)
		datedeb = datetime.datetime.now().strftime("%d-%m-%Y_%H:%M:%S")
		os.mkdir(self.Config.getDataDirname('Exploration Visages') + '/' + name +'_' +  datedeb)

		#SetImageFile = self.SelectionSetIndividu(name)

		SetImageFile = [self.Config.getImageDirname('Exploration Visages/anu/Anubis_02_200110DSC06806.resized.jpg'), self.Config.getImageDirname('Exploration Visages/anu/Anubis_Ref_200123DSC07844.resized.jpg'),self.Config.getImageDirname('Exploration Visages/bar/Barnabe_Ref_200123DSC07867.resized.jpg'),self.Config.getImageDirname('Exploration Visages/bar/Barnabe_04_200123DSC07859.resized.jpg'),self.Config.getImageDirname('Exploration Visages/ces/Cesar_03_200129DSC08442.resized.jpg'),self.Config.getImageDirname('Exploration Visages/ces/Cesar_Ref_200123DSC07860.resized.jpg')]
		random.shuffle (SetImageFile)

		self.tracker.start_recording()
		print("L'expérience a débuté")
		for img in SetImageFile:
			poscercle =self.norm_2_px((xfix,yfix))

			date = datetime.datetime.now().strftime("%d-%m-%Y_%H:%M:%S")
			imgname = os.path.basename(img)
			imgname = os.path.splitext(imgname)[0]
			table = Workbook()
			gazePosSheet = table.active
			gazePosSheet.title = 'gazePos'
			informationsheet = table.create_sheet("Information")
			informationsheet.append(["NomImg","xImage","yImage","xdep","ydep"])
			gazePosSheet.append(["Time", "XOeilDroit", "YOeilDroit","XOeilGauche","YOeilGauche","xRetenu","Yretenu","Etat"])
			informationsheet.append([img,960,702,poscercle[0],poscercle[1]])
			screen = libscreen.Screen()

			#le cercle reste blanc

			screen.draw_circle(colour='white', pos= poscercle, r=30, pw=2, fill=True) 
			self.disp.fill(screen=screen)
			self.disp.show()
			tfix = 0
			cptperte = 0
			while tfix < 250:  

				if self.kb.get_key(keylist= ['space'], flush=False)[0]:
					screen.clear()
					self.disp.fill(screen=screen)
					self.disp.show()
					self.tracker.stop_recording()

					return()

				newTime = libtime.get_time()
				time.sleep(0.005)
				NewTimeStamp, Newgazepos = self.tracker.binocular_sample()

				# si l'individu ne regarde pas le point central

				if (Newgazepos[2][0] < poscercle[0]-100 or Newgazepos[2][0] > poscercle[0] +100) or (Newgazepos[2][1] < poscercle[1]-100 or Newgazepos[2][1] > poscercle[1] + 100) : 
					cptperte += (libtime.get_time() - newTime)
					if cptperte > 300:
						tfix = 0
						screen.draw_circle(colour='white', pos= poscercle, r=30, pw=2, fill=True) 
						self.disp.fill(screen=screen)
						self.disp.show()

				else : # si l'individu regarde le point 
					screen = libscreen.Screen()
					#Le cercle devient vert
					screen.draw_circle(colour=(int(128-(tfix*128/250)),128,0), pos= poscercle, r=30, pw=2, fill=True)
					self.disp.fill(screen=screen)
					self.disp.show()
					tfix += (libtime.get_time() - newTime)
					cptperte = 0




			screen.clear()
			screen.draw_image(image =img,pos=(960,702))
			self.disp.fill(screen=screen)
			self.disp.show()

			tdeb = libtime.get_time()
			oldTimeStamp = 0
			txp,gase = self.tracker.binocular_sample()

			while libtime.get_time()-tdeb < 4000 :
				time.sleep(0.01)
				NewTimeStamp, Newgazepos = self.tracker.binocular_sample()
				if NewTimeStamp != oldTimeStamp :
					t = int(NewTimeStamp - txp) /1000
					etat = self.etat_yeux(Newgazepos[0],Newgazepos[1])
					gazePosSheet.append([t, Newgazepos[0][0],Newgazepos[0][1],Newgazepos[1][0],Newgazepos[1][1],Newgazepos[2][0],Newgazepos[2][1],etat])
					oldTimeStamp = NewTimeStamp

			screen.clear()
			self.disp.fill(screen=screen)
			self.disp.show()
			self.RecSound.play()
			table.save(self.Config.getDataDirname('Exploration Visages') + '/' + name +'_' +  datedeb + '/' + name +'_' + imgname + '_' +  date + '.xls')

			time.sleep(2)
		print("L'expérience est terminée")
		self.tracker.stop_recording()


	def SelectionSetIndividu(self,name):

		''' Permet de chosisir de facon compeltemetn un set d'individu parmi toutes les images possibles '''


		excelname =  self.Config.getImageDirname('Exploration Visages') + '/' + 'dom_training_dyads_selection_final.xlsx'
		document = xlrd.open_workbook(excelname) 
		myFeuille = document.sheet_by_index(0)
		rows = myFeuille.nrows
		for row in range(1,rows):
			if myFeuille.cell_value(rowx=row, colx=0)== name:
				NameSet = [str(myFeuille.cell_value(rowx=row, colx=5)),str(myFeuille.cell_value(rowx=row, colx=6)),str(myFeuille.cell_value(rowx=row, colx=7)),str(myFeuille.cell_value(rowx=row, colx=8)),str(myFeuille.cell_value(rowx=row, colx=9)),str(myFeuille.cell_value(rowx=row, colx=10))]


		ListeFileName = []
		for indName in NameSet:
			ListeNomImg = os.listdir(self.Config.getImageDirname('Exploration Visages') + '/' + indName)
			random.shuffle (ListeNomImg)
			NameSelec = ListeNomImg[0]
			ListeFileName += [self.Config.getImageDirname('Exploration Visages') + '/' + indName + '/' +NameSelec]
		return(ListeFileName)
			
	def etat_yeux(self,PosG,PosD):


		''' 
		Paramètres : 
			Coorodnnée de la position de l'oeil droit et de l'oeil Gauche 
		Description :
			Permet de connaître l'état des yeux, un seul fermé deux ou aucun 
		Sortie :
			0, aucun oeil fermé
			1, Oeil Droit fermé
			2, Oeil gauche fermé
			3, Les deux yeux sont fermés
		'''


		if PosG == (-1,-1) and PosD == (-1,-1):
			return 3
		if PosD == (-1,-1):
			return 1
		if PosG == (-1,-1):
			return 2
		return 0

	### Lancement du script ###





app = experience("Experiences EyeTracker")
